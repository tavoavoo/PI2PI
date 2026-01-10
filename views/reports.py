import customtkinter as ctk
from tkinter import filedialog
import csv
from datetime import datetime, timedelta
import os
import itertools
from collections import defaultdict

# Intentamos importar openpyxl con manejo de error
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass 

class ReportesView(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color="transparent")
        self.c = controller
        self.archivo_binance = None

        # --- HEADER ---
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=20)
        ctk.CTkLabel(header, text="Reportes y Auditoría", font=("Arial", 26, "bold")).pack(side="left")

        # --- AREA DE TRABAJO ---
        main_frame = ctk.CTkFrame(self, fg_color="#1a1a1a", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Instrucciones
        ctk.CTkLabel(main_frame, text="GENERADOR DE REPORTE FISCAL (.xlsx)", font=("Arial", 16, "bold"), text_color="#3498db").pack(pady=(20, 5))
        
        info_text = (
            "Motor Híbrido: Fusiona historial de Binance (CSV) con Operaciones Manuales (DB).\n"
            "Detecta automáticamente: Stock remanente, Ganancia en USDT y Retiros Personales."
        )
        ctk.CTkLabel(main_frame, text=info_text, font=("Arial", 12), text_color="gray", justify="center").pack(pady=5)

        # --- CONTENEDOR DE BOTONES ---
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(pady=15)

        # 1. Cargar CSV
        self.btn_load = ctk.CTkButton(btn_frame, text="📂 CARGAR CSV BINANCE", width=200, height=40, 
                                      font=("Arial", 12, "bold"), command=self.cargar_csv)
        self.btn_load.pack(side="left", padx=10)

        # 2. Carga Manual
        self.btn_manual = ctk.CTkButton(btn_frame, text="➕ NUEVA OP. MANUAL", width=200, height=40, 
                                        font=("Arial", 12, "bold"), fg_color="#8e44ad", hover_color="#732d91",
                                        command=self.abrir_carga_manual) # Usamos la función dedicada
        self.btn_manual.pack(side="left", padx=10)

        self.lbl_file = ctk.CTkLabel(main_frame, text="Ningún archivo seleccionado", font=("Consolas", 12), text_color="#e74c3c")
        self.lbl_file.pack(pady=5)

        # --- FILTRO DE FECHAS ---
        filter_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        filter_frame.pack(pady=10)
        
        ctk.CTkLabel(filter_frame, text="Desde (dd/mm/aaaa):", font=("Arial", 12)).pack(side="left", padx=5)
        self.entry_start = ctk.CTkEntry(filter_frame, width=100, placeholder_text="01/01/2025")
        self.entry_start.pack(side="left", padx=5)
        
        ctk.CTkLabel(filter_frame, text="Hasta (dd/mm/aaaa):", font=("Arial", 12)).pack(side="left", padx=5)
        self.entry_end = ctk.CTkEntry(filter_frame, width=100, placeholder_text="31/01/2026")
        self.entry_end.pack(side="left", padx=5)

        # --- PROCESAR ---
        self.btn_process = ctk.CTkButton(main_frame, text="⚙️ PROCESAR Y GUARDAR EXCEL", width=250, height=50, 
                                         font=("Arial", 13, "bold"), fg_color="#27ae60", hover_color="#219150", 
                                         state="disabled", command=self.procesar_datos)
        self.btn_process.pack(pady=20)

        self.log_box = ctk.CTkTextbox(main_frame, width=600, height=150)
        self.log_box.pack(pady=20)
        self.log("Sistema listo. Esperando archivo...")

    # --- MÉTODO FALTANTE QUE CAUSABA EL ERROR ---
    def abrir_carga_manual(self):
        try:
            self.c.show_view("NuevaOperacionView")
        except AttributeError:
            print("Error: El controlador no tiene show_view. Intentando acceso directo.")
            if hasattr(self.c, 'controller'):
                self.c.controller.show_view("NuevaOperacionView")

    def log(self, text):
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")
        print(text)

    def cargar_csv(self):
        filename = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if filename:
            self.archivo_binance = filename
            self.lbl_file.configure(text=f"Archivo: {os.path.basename(filename)}", text_color="#2ecc71")
            self.btn_process.configure(state="normal")
            self.log(f"Archivo cargado: {filename}")

    # --- UTILIDADES ---
    def clean_decimal(self, valor_str):
        if not valor_str: return 0.0
        clean = "".join(c for c in str(valor_str) if c.isdigit() or c in [',', '.', '-', 'E', 'e', '+'])
        if not clean: return 0.0
        if 'E' in clean.upper(): 
            try:
                val = float(clean.replace(',', '.'))
                if abs(val) < 0.00000001: return 0.0
                return val
            except: pass
        if '.' in clean and ',' in clean:
            last_dot = clean.rfind('.')
            last_comma = clean.rfind(',')
            if last_comma > last_dot: clean = clean.replace('.', '').replace(',', '.') 
            else: clean = clean.replace(',', '') 
        elif ',' in clean: clean = clean.replace(',', '.')
        try: return float(clean)
        except ValueError: return 0.0

    def clean_order_id(self, raw_id):
        if not raw_id: return ""
        s_id = str(raw_id).strip().replace('"', '').replace("'", "")
        if s_id.endswith(".0"): s_id = s_id[:-2]
        return s_id

    def cargar_mapa_bancos_y_personal(self):
        try:
            try:
                self.c.cursor.execute("SELECT order_id, banco, es_personal FROM operaciones")
            except:
                self.c.cursor.execute("SELECT order_id, banco, 0 FROM operaciones")
            
            rows = self.c.cursor.fetchall()
            mapa = {}
            for r in rows:
                oid = str(r[0])
                if oid and oid != "None":
                    banco = r[1]
                    es_personal = True if r[2] == 1 else False
                    mapa[oid] = {'banco': banco, 'es_personal': es_personal}
            self.log(f"✅ Base de Datos conectada. {len(mapa)} ops mapeadas.")
            return mapa
        except Exception as e:
            self.log(f"⚠️ No se pudo leer la base de datos local: {e}")
            return {}

    def procesar_datos(self):
        if not self.archivo_binance: return
        
        try:
            import openpyxl
        except ImportError:
            self.c.show_error("Falta librería", "Necesitas instalar openpyxl.\nEjecuta: pip install openpyxl")
            return

        str_start = self.entry_start.get().strip()
        str_end = self.entry_end.get().strip()
        filter_start = None
        filter_end = None

        try:
            if str_start: filter_start = datetime.strptime(str_start, "%d/%m/%Y")
            if str_end: filter_end = datetime.strptime(str_end, "%d/%m/%Y").replace(hour=23, minute=59, second=59)
        except ValueError:
            self.c.show_error("Error de Fecha", "Formato de fecha inválido. Usa: dd/mm/aaaa")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel File", "*.xlsx")], initialfile="REPORTE_CONTADOR_FINAL.xlsx")
        if not save_path: return

        try:
            self.log("⏳ Iniciando procesamiento híbrido (CSV + Manuales)...")
            
            db_map = self.cargar_mapa_bancos_y_personal()
            fecha_cambio_comision = datetime(2025, 12, 29, 23, 59, 59)
            
            data_por_dia = defaultdict(lambda: {"compras": [], "ventas": []})
            rangos_mensuales = defaultdict(lambda: {"c_usdt": [], "v_usdt": [], "c_ars": [], "v_ars": []})
            lista_personales = [] 
            ids_procesados = set() 
            
            count_compras = 0; count_ventas = 0
            count_skipped_currency = 0; count_skipped_personal = 0

            # --- FASE 1: LECTURA CSV (Binance) ---
            with open(self.archivo_binance, 'r', encoding='utf-8', errors='replace') as f_in:
                sample = f_in.read(2048)
                f_in.seek(0)
                try: dialect = csv.Sniffer().sniff(sample)
                except: dialect = 'excel'
                reader = csv.DictReader(f_in, dialect=dialect)
                if reader.fieldnames:
                    reader.fieldnames = [h.strip().replace('\ufeff', '') for h in reader.fieldnames]
                
                for i, row in enumerate(reader):
                    row_clean = {k.strip(): v for k, v in row.items() if k}
                    if row_clean.get('Status') != 'Completed': continue
                    
                    order_type = row_clean.get('Order Type', '')
                    if order_type not in ['Buy', 'Sell']: continue 

                    fecha_str = row_clean.get('Created Time', '')
                    try: dt_obj = datetime.strptime(fecha_str, "%Y-%m-%d %H:%M:%S")
                    except ValueError: continue

                    dt_real = dt_obj - timedelta(hours=3)
                    if filter_start and dt_real < filter_start: continue
                    if filter_end and dt_real > filter_end: continue

                    fecha_solo_dia = dt_real.strftime("%d/%m/%Y")
                    price_unitario_binance = self.clean_decimal(row_clean.get('Price', '0'))
                    total_pesos = self.clean_decimal(row_clean.get('Total Price', '0'))
                    qty_bruta = self.clean_decimal(row_clean.get('Quantity', '0'))

                    if price_unitario_binance == 0: continue

                    raw_id = row_clean.get('Order Number', '')
                    clean_id = self.clean_order_id(raw_id)
                    ids_procesados.add(clean_id) 

                    counterparty = row_clean.get('Counterparty Nickname', 'N/A')
                    fiat_type = row_clean.get('Fiat Type', '').upper()

                    # 1. FILTRO MONEDA EXTRANJERA
                    if fiat_type != 'ARS':
                        lista_personales.append({
                            "fecha": fecha_solo_dia,
                            "tipo": f"OP. EXTRANJERA ({fiat_type})",
                            "usdt": qty_bruta,
                            "monto_sumable": 0.0
                        })
                        count_skipped_currency += 1
                        continue 

                    # 2. FILTRO PERSONAL
                    es_personal = False
                    banco_detectado = ""
                    if clean_id in db_map:
                        info_db = db_map[clean_id]
                        banco_detectado = info_db.get('banco', '')
                        es_personal = info_db.get('es_personal', False)
                        if banco_detectado == "Por Clasificar": banco_detectado = ""

                    if es_personal:
                        tipo_mov = "GASTO / RETIRO" if order_type == "Buy" else "INGRESO PERSONAL"
                        lista_personales.append({
                            "fecha": fecha_solo_dia,
                            "tipo": tipo_mov,
                            "usdt": qty_bruta,
                            "monto_sumable": total_pesos
                        })
                        count_skipped_personal += 1
                        continue 

                    # 3. LÓGICA COMERCIAL
                    maker_fee = self.clean_decimal(row_clean.get('Maker Fee', ''))
                    taker_fee = self.clean_decimal(row_clean.get('Taker Fee', ''))
                    fee_descontado = 0.0
                    if maker_fee > 0: fee_descontado = maker_fee
                    elif taker_fee > 0: fee_descontado = taker_fee

                    if order_type == 'Buy':
                        if dt_real <= fecha_cambio_comision: factor = 1.0020
                        else: factor = 1.0016
                        precio_ajustado = price_unitario_binance * factor 
                        if fee_descontado == 0: cantidad_cripto_real = total_pesos / precio_ajustado if precio_ajustado > 0 else 0.0
                        else: cantidad_cripto_real = qty_bruta - fee_descontado
                        fila = [fecha_solo_dia, "USDT", cantidad_cripto_real, precio_ajustado, total_pesos, "BINANCE", banco_detectado]
                        data_por_dia[fecha_solo_dia]["compras"].append(fila)
                        count_compras += 1

                    elif order_type == 'Sell':
                        if dt_real <= fecha_cambio_comision: factor_desc = 0.0020
                        else: factor_desc = 0.0016
                        precio_ajustado = price_unitario_binance * (1 - factor_desc)
                        if fee_descontado == 0: cantidad_cripto_real = total_pesos / precio_ajustado if precio_ajustado > 0 else 0.0
                        else: cantidad_cripto_real = qty_bruta + fee_descontado 
                        fila = [fecha_solo_dia, "USDT", cantidad_cripto_real, precio_ajustado, total_pesos, "BINANCE", banco_detectado]
                        data_por_dia[fecha_solo_dia]["ventas"].append(fila)
                        count_ventas += 1

            # --- FASE 2: LECTURA DB (MANUALES) ---
            self.log("🔍 Buscando operaciones manuales en DB...")
            try:
                self.c.cursor.execute("SELECT id, fecha, tipo, monto_ars, monto_usdt, cotizacion, es_personal, order_id, banco FROM operaciones")
                all_ops_db = self.c.cursor.fetchall()
                
                count_manuales = 0
                for op in all_ops_db:
                    oid_db = str(op[7]) if op[7] else ""
                    if oid_db in ids_procesados: continue
                    
                    fecha_db_str = op[1]
                    try:
                        if len(fecha_db_str) > 10: dt_db = datetime.strptime(fecha_db_str, "%Y-%m-%d %H:%M:%S")
                        else: dt_db = datetime.strptime(fecha_db_str, "%Y-%m-%d")
                    except: continue 
                    
                    if filter_start and dt_db < filter_start: continue
                    if filter_end and dt_db > filter_end: continue
                    
                    fecha_solo_dia = dt_db.strftime("%d/%m/%Y")
                    tipo = op[2] 
                    fiat = float(op[3]) if op[3] else 0.0
                    usdt = float(op[4]) if op[4] else 0.0
                    cot = float(op[5]) if op[5] else 0.0
                    es_p = True if op[6] == 1 else False
                    banco = op[8] if op[8] else "Manual"
                    
                    ids_procesados.add(oid_db) 
                    
                    if es_p:
                        # CLASIFICACIÓN SEMÁNTICA PARA MANUALES
                        tipo_mov = "GASTO / RETIRO" 
                        if tipo == "Venta" and usdt > 0: tipo_mov = "RETIRO STOCK (MANUAL)"
                        if tipo == "Compra" and usdt == 0: tipo_mov = "GASTO PERSONAL (MANUAL)"
                        if tipo == "Venta" and usdt == 0: tipo_mov = "INGRESO CAPITAL (MANUAL)"

                        lista_personales.append({
                            "fecha": fecha_solo_dia,
                            "tipo": tipo_mov,
                            "usdt": usdt,
                            "monto_sumable": fiat
                        })
                        count_manuales += 1
                        count_skipped_personal += 1
                    else:
                        row_data = [fecha_solo_dia, "USDT", usdt, cot, fiat, "MANUAL", banco]
                        if tipo == "Compra":
                            data_por_dia[fecha_solo_dia]["compras"].append(row_data)
                            count_compras += 1
                        elif tipo == "Venta":
                            data_por_dia[fecha_solo_dia]["ventas"].append(row_data)
                            count_ventas += 1
                        count_manuales += 1

                if count_manuales > 0: self.log(f"✅ Se integraron {count_manuales} operaciones MANUALES.")
            
            except Exception as e_db: self.log(f"⚠️ Error leyendo manuales: {e_db}")

            if count_skipped_currency > 0: self.log(f"💜 {count_skipped_currency} Ops Extranjeras al Anexo.")
            if count_skipped_personal > 0: self.log(f"💜 {count_skipped_personal} Ops Personales al Anexo.")

            # --- EXCEL ---
            wb = Workbook(); ws = wb.active; ws.title = "Reporte Fiscal"
            font_title = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            font_header = Font(name='Arial', size=10, bold=True)
            font_resumen = Font(name='Arial', size=10, bold=True)
            font_note = Font(name='Arial', size=9, italic=True, color="555555")
            align_center = Alignment(horizontal='center', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')
            fill_green = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            fill_red = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
            fill_blue_light = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            fill_blue_dark = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            fill_purple_light = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
            fmt_number = '#,##0.00'; fmt_accounting = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            
            ws['A1'] = "COMPRA"; ws.merge_cells('A1:G1'); ws['A1'].font = font_title; ws['A1'].alignment = align_center; ws['A1'].fill = fill_green
            ws['H1'] = "VENTA"; ws.merge_cells('H1:N1'); ws['H1'].font = font_title; ws['H1'].alignment = align_center; ws['H1'].fill = fill_red

            headers = ["FECHA", "TIPO DE CRIPTO", "USDT COMPRADO", "PRECIO DE COMPRA", "MONTO EN $", "EXCHANGUE", "MEDIO DE PAGO", "FECHA", "TIPO DE CRIPTO", "USDT VENDIDO", "PRECIO DE VENTA", "MONTO EN $", "EXCHANGUE", "MEDIO DE PAGO"]
            ws.append(headers); 
            for cell in ws[2]: cell.font = font_header; cell.alignment = align_center

            # RENDER DIARIO
            fechas_ordenadas = sorted(data_por_dia.keys(), key=lambda x: datetime.strptime(x, "%d/%m/%Y"))
            row_idx = 3
            
            for fecha in fechas_ordenadas:
                bloque = data_por_dia[fecha]; compras_dia = bloque["compras"]; ventas_dia = bloque["ventas"]
                start_row = row_idx
                rows_count = max(len(compras_dia), len(ventas_dia))
                end_row = start_row + rows_count - 1
                dt_obj = datetime.strptime(fecha, "%d/%m/%Y"); key_mes = dt_obj.strftime("%m/%Y")
                
                if len(compras_dia) > 0:
                    rangos_mensuales[key_mes]["c_usdt"].append(f"C{start_row}:C{start_row + len(compras_dia) - 1}")
                    rangos_mensuales[key_mes]["c_ars"].append(f"E{start_row}:E{start_row + len(compras_dia) - 1}")
                if len(ventas_dia) > 0:
                    rangos_mensuales[key_mes]["v_usdt"].append(f"J{start_row}:J{start_row + len(ventas_dia) - 1}")
                    rangos_mensuales[key_mes]["v_ars"].append(f"L{start_row}:L{start_row + len(ventas_dia) - 1}")

                for c_data, v_data in itertools.zip_longest(compras_dia, ventas_dia, fillvalue=None):
                    if c_data:
                        ws.cell(row=row_idx, column=1, value=c_data[0]).alignment = align_center
                        ws.cell(row=row_idx, column=2, value=c_data[1]).alignment = align_center
                        ws.cell(row=row_idx, column=3, value=c_data[2]).number_format = fmt_number; ws.cell(row=row_idx, column=3).alignment = align_right
                        ws.cell(row=row_idx, column=4, value=c_data[3]).number_format = fmt_number; ws.cell(row=row_idx, column=4).alignment = align_right
                        ws.cell(row=row_idx, column=5, value=c_data[4]).number_format = fmt_accounting; ws.cell(row=row_idx, column=5).alignment = align_right
                        ws.cell(row=row_idx, column=6, value=c_data[5]).alignment = align_center
                        ws.cell(row=row_idx, column=7, value=c_data[6]).alignment = align_center
                    else: ws.cell(row=row_idx, column=1, value=fecha).alignment = align_center

                    if v_data:
                        ws.cell(row=row_idx, column=8, value=v_data[0]).alignment = align_center
                        ws.cell(row=row_idx, column=9, value=v_data[1]).alignment = align_center
                        ws.cell(row=row_idx, column=10, value=v_data[2]).number_format = fmt_number; ws.cell(row=row_idx, column=10).alignment = align_right
                        ws.cell(row=row_idx, column=11, value=v_data[3]).number_format = fmt_number; ws.cell(row=row_idx, column=11).alignment = align_right
                        ws.cell(row=row_idx, column=12, value=v_data[4]).number_format = fmt_accounting; ws.cell(row=row_idx, column=12).alignment = align_right
                        ws.cell(row=row_idx, column=13, value=v_data[5]).alignment = align_center
                        ws.cell(row=row_idx, column=14, value=v_data[6]).alignment = align_center
                    else: ws.cell(row=row_idx, column=8, value=fecha).alignment = align_center
                    row_idx += 1
                
                # SEPARADOR VISUAL
                cell_res = ws.cell(row=row_idx, column=1, value="")
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=14)
                cell_res.fill = fill_blue_light; row_idx += 1

            # --- TABLAS RESUMEN ---
            border_unified = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            font_gd = Font(name='Arial', size=11, bold=True)
            meses_ordenados = sorted(rangos_mensuales.keys(), key=lambda x: datetime.strptime(x, "%m/%Y"))
            def make_sum(r): return f"SUM({','.join(r)})" if r else "0"
            def sum_refs(r): return f"=SUM({','.join(r)})" if r else "=0"

            # 1. USDT
            row_idx += 2
            ws.cell(row=row_idx, column=1, value="RESUMEN OPERATIVO (USDT)").font = Font(name='Arial', size=12, bold=True)
            row_idx += 1
            headers_usdt = ["MES", "COMPRAS (USDT)", "VENTAS (USDT)", "STOCK (BALANCE)", "GANANCIA EST. (USDT)"]
            for i, h in enumerate(headers_usdt, 1):
                cell = ws.cell(row=row_idx, column=i, value=h)
                cell.fill = fill_blue_dark; cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                cell.alignment = align_center; cell.border = border_unified
            row_idx += 1
            
            total_c_usdt = []; total_v_usdt = []; total_g_usdt = []
            for mes in meses_ordenados:
                rgs = rangos_mensuales[mes]
                ws.cell(row=row_idx, column=1, value=mes).border = border_unified; ws.cell(row=row_idx, column=1).alignment = align_center
                c_buy = ws.cell(row=row_idx, column=2, value="="+make_sum(rgs["c_usdt"])); c_buy.number_format=fmt_number; c_buy.border=border_unified; total_c_usdt.append(f"B{row_idx}")
                c_sell = ws.cell(row=row_idx, column=3, value="="+make_sum(rgs["v_usdt"])); c_sell.number_format=fmt_number; c_sell.border=border_unified; total_v_usdt.append(f"C{row_idx}")
                ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").number_format=fmt_number; ws.cell(row=row_idx, column=4).border=border_unified
                sum_v_ars = make_sum(rgs["v_ars"]); sum_c_ars = make_sum(rgs["c_ars"]); sum_v_usdt = f"C{row_idx}" 
                form_profit = f'=IF({sum_v_usdt}<>0, ({sum_v_ars}-{sum_c_ars})/({sum_v_ars}/{sum_v_usdt}), 0)'
                c_prof = ws.cell(row=row_idx, column=5, value=form_profit); c_prof.number_format=fmt_number; c_prof.border=border_unified; total_g_usdt.append(f"E{row_idx}")
                row_idx += 1

            ws.cell(row=row_idx, column=1, value="TOTAL").font=font_gd; ws.cell(row=row_idx, column=1).border=border_unified
            ws.cell(row=row_idx, column=2, value=sum_refs(total_c_usdt)).font=font_gd; ws.cell(row=row_idx, column=2).border=border_unified; ws.cell(row=row_idx, column=2).number_format=fmt_number
            ws.cell(row=row_idx, column=3, value=sum_refs(total_v_usdt)).font=font_gd; ws.cell(row=row_idx, column=3).border=border_unified; ws.cell(row=row_idx, column=3).number_format=fmt_number
            ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").font=font_gd; ws.cell(row=row_idx, column=4).border=border_unified; ws.cell(row=row_idx, column=4).number_format=fmt_number
            ws.cell(row=row_idx, column=5, value=sum_refs(total_g_usdt)).font=font_gd; ws.cell(row=row_idx, column=5).border=border_unified; ws.cell(row=row_idx, column=5).number_format=fmt_number

            # 2. ARS
            row_idx += 3
            ws.cell(row=row_idx, column=1, value="RESUMEN FINANCIERO (PESOS)").font = Font(name='Arial', size=12, bold=True)
            row_idx += 1
            headers_ars = ["MES", "COMPRAS / GASTOS ($)", "VENTAS / INGRESOS ($)", "RESULTADO CAJA ($)"]
            for i, h in enumerate(headers_ars, 1):
                cell = ws.cell(row=row_idx, column=i, value=h)
                cell.fill = fill_blue_dark; cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                cell.alignment = align_center; cell.border = border_unified
            row_idx += 1
            
            total_c_ars = []; total_v_ars = []
            for mes in meses_ordenados:
                rgs = rangos_mensuales[mes]
                ws.cell(row=row_idx, column=1, value=mes).border = border_unified; ws.cell(row=row_idx, column=1).alignment = align_center
                c_cars = ws.cell(row=row_idx, column=2, value="="+make_sum(rgs["c_ars"])); c_cars.number_format=fmt_accounting; c_cars.border=border_unified; total_c_ars.append(f"B{row_idx}")
                c_vars = ws.cell(row=row_idx, column=3, value="="+make_sum(rgs["v_ars"])); c_vars.number_format=fmt_accounting; c_vars.border=border_unified; total_v_ars.append(f"C{row_idx}")
                ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").number_format=fmt_accounting; ws.cell(row=row_idx, column=4).border=border_unified
                row_idx += 1

            ws.cell(row=row_idx, column=1, value="TOTAL").font=font_gd; ws.cell(row=row_idx, column=1).border=border_unified
            ws.cell(row=row_idx, column=2, value=sum_refs(total_c_ars)).font=font_gd; ws.cell(row=row_idx, column=2).border=border_unified; ws.cell(row=row_idx, column=2).number_format=fmt_accounting
            ws.cell(row=row_idx, column=3, value=sum_refs(total_v_ars)).font=font_gd; ws.cell(row=row_idx, column=3).border=border_unified; ws.cell(row=row_idx, column=3).number_format=fmt_accounting
            ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").font=font_gd; ws.cell(row=row_idx, column=4).border=border_unified; ws.cell(row=row_idx, column=4).number_format=fmt_accounting

            # ANEXO
            if lista_personales:
                row_idx += 3
                ws.cell(row=row_idx, column=1, value="ANEXO: MOVIMIENTOS PERSONALES Y RETIROS").font = Font(name='Arial', size=12, bold=True)
                row_idx += 1
                headers_p = ["FECHA", "TIPO", "USDT RETIRADO"]
                for i, h in enumerate(headers_p, 1):
                    c = ws.cell(row=row_idx, column=i, value=h)
                    c.fill = fill_purple_light; c.font = Font(name='Arial', size=10, bold=True); c.border = border_unified; c.alignment = align_center
                row_idx += 1
                start_p = row_idx
                for item in lista_personales:
                    ws.cell(row=row_idx, column=1, value=item['fecha']).border = border_unified; ws.cell(row=row_idx, column=1).alignment = align_center
                    ws.cell(row=row_idx, column=2, value=item['tipo']).border = border_unified; ws.cell(row=row_idx, column=2).alignment = align_center
                    ws.cell(row=row_idx, column=3, value=item['usdt']).border = border_unified; ws.cell(row=row_idx, column=3).number_format = fmt_number; ws.cell(row=row_idx, column=3).alignment = align_right
                    ws.cell(row=row_idx, column=4, value=item['monto_sumable']) 
                    row_idx += 1
                end_p = row_idx - 1
                ws.cell(row=row_idx, column=2, value="TOTAL:").font = font_gd; ws.cell(row=row_idx, column=2).alignment = align_right
                ws.cell(row=row_idx, column=3, value=f"=SUM(C{start_p}:C{end_p})").font = font_gd; ws.cell(row=row_idx, column=3).number_format = fmt_number
                ws.cell(row=row_idx, column=4, value=f"=SUM(D{start_p}:D{end_p})").number_format = fmt_accounting; ws.column_dimensions['D'].hidden = True 

            # Ancho columnas
            for i, col in enumerate(ws.columns, 1):
                col_letter = get_column_letter(i)
                if i in [1, 8]: ws.column_dimensions[col_letter].width = 15; continue
                if i in [4, 5, 11, 12]: ws.column_dimensions[col_letter].width = 22; continue
                if i in [7, 14]: ws.column_dimensions[col_letter].width = 20; continue
                if row_idx > 10 and i in [4, 5, 7, 8]: ws.column_dimensions[col_letter].width = 24; continue 
                if lista_personales and i == 2: ws.column_dimensions[col_letter].width = 30; continue 
                max_len = 0
                for cell in col:
                    try:
                        if cell.value: val_len = len(str(cell.value))
                        else: val_len = 0
                        if val_len > max_len: max_len = val_len
                    except: pass
                final_w = (max_len + 4) if max_len > 0 else 12
                if final_w > 30: final_w = 30
                ws.column_dimensions[col_letter].width = final_w

            wb.save(save_path)
            self.log(f"✅ ¡Éxito! Reporte ARS Final.")
            msg_p = f"\nSe listaron {len(lista_personales)} mov. personales en el Anexo." if lista_personales else ""
            self.c.show_info("Listo", f"Reporte generado.{msg_p}")

        except Exception as e:
            self.log(f"❌ Error crítico: {str(e)}")
            self.c.show_error("Error", str(e))